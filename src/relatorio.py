"""Módulo para geração de relatórios Excel."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.config import Config
from src.config_manager import ConfigManager
from src.models import (
    ResultadoComparacao,
)


class GeradorRelatorio:
    """Classe para geração de relatórios Excel."""

    def __init__(self):
        self.resultado: ResultadoComparacao | None = None
        self.arquivo_saida: Path | None = None
        self.nome_arquivo: str = Config.RELATORIO_NOME_PADRAO
        self.pasta_destino: Path = Config.REPORTS_DIR

    def carregar_resultado(self, resultado: ResultadoComparacao) -> None:
        """Carrega o resultado da comparação."""
        self.resultado = resultado
        self.resultado.calcular_totais()

    def definir_nome_arquivo(self, nome: str) -> None:
        """Define o nome do arquivo de saída."""
        if not nome.endswith('.xlsx'):
            nome += '.xlsx'
        self.nome_arquivo = nome

    def definir_pasta_destino(self, pasta: str) -> None:
        """Define a pasta de destino."""
        self.pasta_destino = Path(pasta)
        self.pasta_destino.mkdir(parents=True, exist_ok=True)

    def gerar(self) -> Path:
        """Gera o relatório Excel."""
        if not self.resultado:
            raise ValueError("Nenhum resultado carregado. Execute carregar_resultado() primeiro.")

        # Carrega configurações
        config_manager = ConfigManager()
        colunas_selecionadas = config_manager.get_colunas_selecionadas()
        abas_selecionadas = config_manager.get_abas_selecionadas()
        ordenacao = config_manager.get_ordenacao()

        # Aplica ordenação
        self._aplicar_ordenacao(ordenacao)

        # Cria o caminho completo
        self.arquivo_saida = self.pasta_destino / self.nome_arquivo

        # Verifica se arquivo existe e gera nome alternativo
        if self.arquivo_saida.exists():
            contador = 1
            while True:
                novo_nome = self.arquivo_saida.stem + f"_{contador}" + self.arquivo_saida.suffix
                novo_caminho = self.pasta_destino / novo_nome
                if not novo_caminho.exists():
                    self.arquivo_saida = novo_caminho
                    break
                contador += 1

        # Cria o Excel com múltiplas abas
        with pd.ExcelWriter(self.arquivo_saida, engine='openpyxl') as writer:
            # Aba: Divergências
            if "Divergências" in abas_selecionadas:
                self._criar_aba_divergencias(writer, colunas_selecionadas)

            # Aba: Apenas na A
            if "Apenas na A" in abas_selecionadas:
                self._criar_aba_apenas_a(writer, colunas_selecionadas)

            # Aba: Apenas na B
            if "Apenas na B" in abas_selecionadas:
                self._criar_aba_apenas_b(writer, colunas_selecionadas)

            # Aba: Iguais
            if "Iguais" in abas_selecionadas:
                self._criar_aba_iguais(writer, colunas_selecionadas)

            # Aba: Erros
            if "Erros" in abas_selecionadas:
                self._criar_aba_erros(writer, colunas_selecionadas)

            # Aba: Resumo (sempre incluída)
            self._criar_aba_resumo(writer)

        # Aplica formatação
        self._aplicar_formatacao(self.arquivo_saida)

        return self.arquivo_saida

    def _aplicar_ordenacao(self, criterio: str) -> None:
        """Aplica ordenação aos dados conforme critério."""
        if criterio == "cpf":
            key = lambda x: x.cpf
        elif criterio == "matricula":
            key = lambda x: x.matricula
        elif criterio == "nome":
            key = lambda x: x.nome or ""
        elif criterio == "data":
            key = lambda x: x.data_admissao
        else:
            return

        self.resultado.apenas_a.sort(key=key)
        self.resultado.apenas_b.sort(key=key)
        self.resultado.iguais.sort(key=key)
        self.resultado.divergencias.sort(key=lambda x: x.cpf if criterio == "cpf" else x.matricula)

    def _criar_aba_divergencias(self, writer: pd.ExcelWriter, colunas: list) -> None:
        """
        Cria a aba de divergências.

        Args:
            writer: ExcelWriter para escrever os dados
            colunas: Lista de colunas selecionadas para incluir
        """
        if not self.resultado.divergencias:
            df = pd.DataFrame({'Mensagem': ['Nenhuma divergência encontrada']})
            df.to_excel(writer, sheet_name='Divergências', index=False)
            return

        dados = []
        for div in self.resultado.divergencias:
            linha = {}
            
            # Colunas obrigatórias (sempre incluídas)
            if True:  # CPF sempre incluído
                linha['CPF'] = div.cpf_mascarado()
            
            if True:  # Matrícula sempre incluída
                linha['Matrícula'] = div.matricula
            
            # Colunas opcionais (conforme configuração)
            if 'Nome' in colunas:
                linha['Nome A'] = div.nome_a or ''
                linha['Nome B'] = div.nome_b or ''
            
            if 'Data de Admissão' in colunas:
                linha['Data A'] = div.data_a_brasil()
                linha['Data B'] = div.data_b_brasil()
            
            # Diferença sempre incluída
            linha['Diferença (dias)'] = div.diferenca_dias
            
            # Colunas extras (dados extras das planilhas)
            if div.dados_extras_a:
                for key, value in div.dados_extras_a.items():
                    if key in colunas:
                        linha[f'A_{key}'] = value
            
            if div.dados_extras_b:
                for key, value in div.dados_extras_b.items():
                    if key in colunas:
                        linha[f'B_{key}'] = value
            
            dados.append(linha)

        # Cria DataFrame com as colunas na ordem desejada
        colunas_ordem = []
        if True:
            colunas_ordem.append('CPF')
        if True:
            colunas_ordem.append('Matrícula')
        if 'Nome' in colunas:
            colunas_ordem.extend(['Nome A', 'Nome B'])
        if 'Data de Admissão' in colunas:
            colunas_ordem.extend(['Data A', 'Data B'])
        colunas_ordem.append('Diferença (dias)')
        
        # Adiciona colunas extras que estão em colunas
        if dados:
            for key in dados[0]:
                if key not in colunas_ordem:
                    colunas_ordem.append(key)
        
        df = pd.DataFrame(dados)
        
        # Reordena colunas
        df = df[colunas_ordem] if all(c in df.columns for c in colunas_ordem) else df
        
        df.to_excel(writer, sheet_name='Divergências', index=False)

    def _criar_aba_apenas_a(self, writer: pd.ExcelWriter) -> None:
        """Cria a aba de registros apenas na Planilha A."""
        if not self.resultado.apenas_a:
            df = pd.DataFrame({'Mensagem': ['Nenhum registro apenas na Planilha A']})
            df.to_excel(writer, sheet_name='Apenas na A', index=False)
            return

        dados = []
        for reg in self.resultado.apenas_a:
            dados.append({
                'CPF': reg.cpf_mascarado(),
                'Matrícula': reg.matricula,
                'Nome': reg.nome or '',
                'Data de Admissão': reg.data_admissao_brasil(),
                'Status': 'Não encontrado na Planilha B',
            })

        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Apenas na A', index=False)

    def _criar_aba_apenas_b(self, writer: pd.ExcelWriter) -> None:
        """Cria a aba de registros apenas na Planilha B."""
        if not self.resultado.apenas_b:
            df = pd.DataFrame({'Mensagem': ['Nenhum registro apenas na Planilha B']})
            df.to_excel(writer, sheet_name='Apenas na B', index=False)
            return

        dados = []
        for reg in self.resultado.apenas_b:
            dados.append({
                'CPF': reg.cpf_mascarado(),
                'Matrícula': reg.matricula,
                'Nome': reg.nome or '',
                'Data de Admissão': reg.data_admissao_brasil(),
                'Status': 'Não encontrado na Planilha A',
            })

        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Apenas na B', index=False)

    def _criar_aba_iguais(self, writer: pd.ExcelWriter) -> None:
        """Cria a aba de registros idênticos."""
        if not self.resultado.iguais:
            df = pd.DataFrame({'Mensagem': ['Nenhum registro idêntico encontrado']})
            df.to_excel(writer, sheet_name='Iguais', index=False)
            return

        dados = []
        for reg in self.resultado.iguais:
            dados.append({
                'CPF': reg.cpf_mascarado(),
                'Matrícula': reg.matricula,
                'Nome': reg.nome or '',
                'Data de Admissão': reg.data_admissao_brasil(),
            })

        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Iguais', index=False)

    def _criar_aba_erros(self, writer: pd.ExcelWriter) -> None:
        """Cria a aba de erros."""
        if not self.resultado.erros:
            df = pd.DataFrame({'Mensagem': ['Nenhum erro encontrado']})
            df.to_excel(writer, sheet_name='Erros', index=False)
            return

        dados = []
        for erro in self.resultado.erros:
            dados.append({
                'Linha': erro.linha,
                'Planilha': erro.planilha,
                'CPF': erro.cpf_mascarado() if erro.cpf else '',
                'Nome': erro.nome or "",
                'Matrícula': erro.matricula or '',
                'Data': erro.data_brasil() if erro.data else '',
                'Erro': erro.erro,
            })

        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Erros', index=False)

    def _criar_aba_resumo(self, writer: pd.ExcelWriter) -> None:
        """Cria a aba de resumo."""
        dados = {
            'Métrica': [
                'Data Processamento',
                'Planilha A',
                'Planilha B',
                'Total Registros A',
                'Total Registros B',
                'Total Divergências',
                'Total Apenas na A',
                'Total Apenas na B',
                'Total Iguais',
                'Total Erros',
                'Taxa de Divergência (%)',
            ],
            'Valor': [
                datetime.now().strftime('%d/%m/%Y %H:%M:%S'),  # noqa: DTZ005
                self.resultado.total_registros_a,
                self.resultado.total_registros_b,
                self.resultado.total_registros_a,
                self.resultado.total_registros_b,
                self.resultado.total_divergencias,
                self.resultado.total_apenas_a,
                self.resultado.total_apenas_b,
                self.resultado.total_iguais,
                self.resultado.total_erros,
                f"{(self.resultado.total_divergencias / max(self.resultado.total_registros_a, 1) * 100):.2f}%",
            ]
        }

        df = pd.DataFrame(dados)
        df.to_excel(writer, sheet_name='Resumo', index=False)

    def _aplicar_formatacao(self, caminho: Path) -> None:
        """Aplica formatação visual ao Excel."""
        try:
            wb = load_workbook(caminho)

            # Estilos
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )

            # Aplica em cada aba
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]

                # Cabeçalho
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # Ajusta largura das colunas
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            max_length = max(max_length, len(str(cell.value)))
                        except:  # noqa: E722, S110
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width

            wb.save(caminho)

        except Exception as e:  # noqa: BLE001
            print(f"⚠️ Erro ao aplicar formatação: {e}")